"""관내출장 결재내역을 사람별로 집계해 Excel로 저장하는 도구."""

from __future__ import annotations

import re
import sys
import tempfile
from calendar import monthrange
from collections import defaultdict
from copy import copy
from pathlib import Path
from typing import Any


PERSON_ALIASES = ("성명", "출장자", "신청자", "이름", "사원명", "사용자")
DEPARTMENT_ALIASES = ("소속", "부서")
RANK_ALIASES = ("직급", "직위")
AMOUNT_ALIASES = ("합계", "지급액", "금액", "출장비", "여비", "총액")
COST_ALIASES = ("일비", "식비", "숙박료", "교통비", "운임", "배차료")
IGNORE_AMOUNT_WORDS = ("일수", "거리", "번호", "순번", "코드", "연번")
TIME_ALIASES = ("출장시간", "출장 시간", "시간", "출발시간", "출발 시간")
START_TIME_ALIASES = ("출발시간", "출발 시간", "출발시각", "출발 시각", "출발")
END_TIME_ALIASES = ("도착시간", "도착 시간", "도착시각", "도착 시각", "도착")
LOCATION_ALIASES = ("근무지내", "근무지", "출장구분", "출장 유형", "출장유형", "구분")
DAY_ALIASES = ("출장일수", "총일수", "일수", "기간")
VEHICLE_ALIASES = ("차량사용여부", "차량 사용 여부", "차량사용", "차량 사용", "차량")
SHORT_TRIP_AMOUNT = 10_000
LONG_TRIP_AMOUNT = 20_000
VEHICLE_DEDUCTION = 10_000
OUTSIDE_DAILY_AMOUNT = 25_000
OUTSIDE_MEAL_AMOUNT = 25_000
OUTSIDE_VEHICLE_DEDUCTION = 12_500
EXTERNAL_CITY_NAMES = ("수원", "시흥", "세종", "서울", "용인", "안산", "오산", "평택", "성남", "안양", "군포", "의왕", "과천", "광명", "부천", "인천")


def clean_text(value: Any) -> str:
    """셀 값에서 줄바꿈과 불필요한 공백을 제거한다."""
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def clean_person_name(value: Any) -> str:
    """성명 아래에 붙은 사번, 숫자, 영문을 제거한다."""
    text = clean_text(value).split("(", 1)[0]
    korean_parts = re.findall(r"[가-힣]+", text)
    return "".join(korean_parts) if korean_parts else re.sub(r"[A-Za-z]", "", text).strip()


def number(value: Any) -> float:
    """원화 표기, 쉼표, 괄호 음수를 포함한 값을 숫자로 바꾼다."""
    text = clean_text(value).replace(",", "").replace("원", "")
    if not text or text in {"-", "—"}:
        return 0.0
    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return 0.0
    result = float(match.group())
    return -result if negative else result


def header_score(header: str, aliases: tuple[str, ...]) -> int:
    normalized = clean_text(header).replace(" ", "")
    return max((len(alias) for alias in aliases if alias in normalized), default=0)


def choose_column(headers: list[str], aliases: tuple[str, ...]) -> int | None:
    scores = [header_score(header, aliases) for header in headers]
    best = max(scores, default=0)
    return scores.index(best) if best else None


def duration_hours(value: Any) -> float | None:
    """시간 범위 또는 '3시간 30분' 표기를 시간 단위로 변환한다."""
    text = clean_text(value).replace("：", ":").replace("．", ".")
    range_match = re.search(r"(\d{1,2}):(\d{2})\s*[-~∼至에서]\s*(\d{1,2}):(\d{2})", text)
    if range_match:
        start = int(range_match.group(1)) * 60 + int(range_match.group(2))
        end = int(range_match.group(3)) * 60 + int(range_match.group(4))
        if end < start:
            end += 24 * 60
        return (end - start) / 60
    clock_matches = re.findall(r"(\d{1,2})\s*(?::|시|\.\s*)(\d{1,2})?\s*분?", text)
    if len(clock_matches) >= 2:
        start = int(clock_matches[0][0]) * 60 + int(clock_matches[0][1] or 0)
        end = int(clock_matches[1][0]) * 60 + int(clock_matches[1][1] or 0)
        if end < start:
            end += 24 * 60
        return (end - start) / 60
    if len(clock_matches) == 1 and ":" in text:
        return int(clock_matches[0][0]) + int(clock_matches[0][1] or 0) / 60
    hour_match = re.search(r"(\d+(?:\.\d+)?)\s*시간", text)
    minute_match = re.search(r"(\d+)\s*분", text)
    if hour_match:
        return float(hour_match.group(1)) + (int(minute_match.group(1)) / 60 if minute_match else 0)
    day_match = re.fullmatch(r"\s*(\d+)\s*일\s*", text)
    if day_match:
        return float(day_match.group(1)) * 24
    if minute_match:
        return int(minute_match.group(1)) / 60
    plain_number = re.fullmatch(r"\s*(\d+(?:\.\d+)?)\s*", text)
    return float(plain_number.group(1)) if plain_number else None


def vehicle_used(value: Any) -> bool:
    text = clean_text(value).replace(" ", "")
    if not text or any(word in text for word in ("미사용", "안씀", "없음", "무")):
        return False
    return any(word in text for word in ("사용", "이용", "유", "Y", "예", "있음", "관용차"))


def display_rank(value: Any) -> str:
    text = clean_text(value)
    normalized = text.replace(" ", "")
    replacements = (
        ("지방시설사무관", "시설5급"), ("시설사무관", "시설5급"),
        ("지방행정사무관", "행정5급"), ("행정사무관", "행정5급"),
        ("지방시설주사보", "시설7급"), ("시설주사보", "시설7급"),
        ("지방행정주사보", "행정7급"), ("행정주사보", "행정7급"),
        ("지방시설주사", "시설6급"), ("시설주사", "시설6급"),
        ("지방행정주사", "행정6급"), ("행정주사", "행정6급"),
        ("지방시설서기보", "시설9급"), ("시설서기보", "시설9급"),
        ("지방행정서기보", "행정9급"), ("행정서기보", "행정9급"),
        ("지방시설서기", "시설8급"), ("시설서기", "시설8급"),
        ("지방행정서기", "행정8급"), ("행정서기", "행정8급"),
    )
    for source, target in replacements:
        if source in normalized:
            return target
    return text


def extract_department_and_rank(value: Any) -> tuple[str, str]:
    """소속/직급이 합쳐진 PDF 셀에서 소속과 직급을 분리한다."""
    text = clean_text(value).replace(" ", "")
    department_part = text.split("실", 1)[-1] if "실" in text else text
    department_match = re.search(r"([가-힣]+과)", department_part)
    department = department_match.group(1) if department_match else ""
    return department, display_rank(text)


def outside_destination(value: Any) -> bool:
    """출장지가 화성시 밖의 도시인지 판정한다."""
    text = clean_text(value).replace(" ", "")
    if not text:
        return False
    if any(city in text for city in EXTERNAL_CITY_NAMES):
        return True
    if "화성" in text:
        return False
    return any(match != "화성시" for match in re.findall(r"[가-힣]{2,}시", text))


def month_day_range(detail: list[dict[str, Any]]) -> str:
    dates = [
        (int(year), int(month), int(day))
        for row in detail
        for year, month, day in re.findall(r"(\d{4})[-./년]\s*(\d{1,2})[-./월]\s*(\d{1,2})", row.get("출장기간", ""))
    ]
    if not dates:
        return ""
    year, month, _ = min(dates)
    return f"{month}.1~{month}.{monthrange(year, month)[1]}"


def trip_amount(hours: float | None, location: Any, vehicle: Any) -> int:
    if hours is None:
        raise ValueError("출장시간을 읽지 못했습니다. 시간 범위를 '09:00~13:00'처럼 표시해 주세요.")
    if hours < 4:
        amount = SHORT_TRIP_AMOUNT
    else:
        amount = LONG_TRIP_AMOUNT
    if vehicle_used(vehicle):
        amount -= VEHICLE_DEDUCTION
    return amount


def trip_days(value: Any) -> int:
    match = re.search(r"(\d+)\s*일", clean_text(value))
    return max(int(match.group(1)), 1) if match else 1


def summarize_trip_rows(headers: list[str], rows: list[list[Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    person_index = choose_column(headers, PERSON_ALIASES)
    time_index = choose_column(headers, TIME_ALIASES)
    start_index = choose_column(headers, START_TIME_ALIASES)
    end_index = choose_column(headers, END_TIME_ALIASES)
    vehicle_index = choose_column(headers, VEHICLE_ALIASES)
    location_index = choose_column(headers, LOCATION_ALIASES)
    destination_index = next((index for index, header in enumerate(headers) if "출장지" in clean_text(header)), None)
    day_index = choose_column(headers, DAY_ALIASES)
    department_index = choose_column(headers, DEPARTMENT_ALIASES)
    rank_index = choose_column(headers, RANK_ALIASES)
    if time_index is None and start_index is not None and end_index is not None:
        time_index = start_index
    if person_index is None or time_index is None or vehicle_index is None:
        raise ValueError("PDF에서 성명, 출장시간, 차량사용여부 열을 찾지 못했습니다.")
    detail: list[dict[str, Any]] = []
    totals: dict[str, dict[str, Any]] = defaultdict(lambda: {"소속": "", "직급": "", "관외 출장지": [], "출장 횟수": 0, "관내 출장 수": 0, "관외 총일수": 0, "관내 차량 수": 0, "관외 차량 수": 0, "4시간 미만 수": 0, "총 출장비": 0})
    previous_person = ""
    for row_number, row in enumerate(rows, start=1):
        values = list(row) + [""] * max(0, len(headers) - len(row))
        raw_person = clean_person_name(values[person_index])
        time_text = clean_text(values[time_index])
        if start_index is not None and end_index is not None and start_index != end_index:
            time_text = f"{values[start_index]}~{values[end_index]}"
        location = values[location_index] if location_index is not None else "근무지내"
        destination = values[destination_index] if destination_index is not None else ""
        destination_text = clean_text(destination)
        clear_external_destination = outside_destination(destination_text)
        outside = "지외" in clean_text(location) or "관외" in clean_text(location) or clear_external_destination
        normalized_person = raw_person.replace(" ", "")
        if normalized_person in {"성명", "소계", "합계", "총계"}:
            continue
        if not raw_person and not time_text:
            continue
        if duration_hours(time_text) is None and not outside:
            continue
        person = raw_person or previous_person
        if not person:
            continue
        previous_person = person
        hours = duration_hours(time_text)
        vehicle = values[vehicle_index]
        days = trip_days(time_text) if outside else (trip_days(values[day_index]) if day_index is not None else 1)
        using_vehicle = vehicle_used(vehicle)
        amount = (days * OUTSIDE_DAILY_AMOUNT + days * OUTSIDE_MEAL_AMOUNT - (days * OUTSIDE_VEHICLE_DEDUCTION if using_vehicle else 0)) if outside else trip_amount(hours, location, vehicle)
        department, rank = extract_department_and_rank(values[department_index]) if department_index is not None else ("", "")
        detail.append({
            "성명": person,
            "소속": department or (clean_text(values[department_index]) if department_index is not None else ""),
            "직급": display_rank(values[rank_index]) if rank_index is not None else rank,
            "출장시간": time_text,
            "출장기간": clean_text(values[3]) if len(values) > 3 else "",
            "근무지": "근무지외" if outside else "근무지내",
            "관외여부": outside,
            "출장일수": days,
            "차량사용여부": clean_text(vehicle) or "미사용",
            "적용금액": amount,
        })
        if not totals[person]["소속"]:
            totals[person]["소속"] = department or (clean_text(values[department_index]) if department_index is not None else "")
        if not totals[person]["직급"]:
            totals[person]["직급"] = display_rank(values[rank_index]) if rank_index is not None else rank
        if clear_external_destination and destination_text not in totals[person]["관외 출장지"]:
            totals[person]["관외 출장지"].append(destination_text)
        totals[person]["출장 횟수"] += 1
        if outside:
            totals[person]["관외 총일수"] += days
            totals[person]["관외 차량 수"] += days if using_vehicle else 0
        else:
            totals[person]["관내 출장 수"] += 1
            totals[person]["관내 차량 수"] += int(using_vehicle)
            totals[person]["4시간 미만 수"] += int((hours or 0) < 4)
        totals[person]["총 출장비"] += amount
    summary = [
        {"성명": person, **data}
        for person, data in sorted(totals.items())
    ]
    for person in summary:
        person["비고"] = "관외: " + ", ".join(person.pop("관외 출장지")) if person.get("관외 출장지") else ""
    return detail, summary


def ocr_pdf_rows(path: Path) -> tuple[list[str], list[list[str]]]:
    """표 추출이 안 되는 스캔 PDF를 OCR로 읽는다. 한국어 Tesseract가 필요하다."""
    try:
        import fitz
        import pytesseract
        from PIL import Image
    except ImportError as error:
        raise ValueError("스캔 PDF에는 pymupdf, pytesseract, pillow가 필요합니다.") from error
    rows: list[list[str]] = []
    time_pattern = re.compile(r"\d{1,2}:\d{2}\s*[-~∼]\s*\d{1,2}:\d{2}|\d+(?:\.\d+)?\s*시간(?:\s*\d+\s*분)?")
    skip_words = {"성명", "출장자", "출장시간", "근무지", "차량", "사용여부", "합계", "총계"}
    document = fitz.open(path)
    try:
        for page in document:
            pixels = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
            image = Image.frombytes("RGB", [pixels.width, pixels.height], pixels.samples)
            text = pytesseract.image_to_string(image, lang="kor+eng")
            for line in text.splitlines():
                line = clean_text(line)
                time_match = time_pattern.search(line)
                if not time_match:
                    continue
                name_candidates = [word for word in re.findall(r"[가-힣]{2,5}", line) if word not in skip_words]
                if not name_candidates:
                    continue
                vehicle = "사용" if re.search(r"차량\s*(사용|이용)|관용차|\bY\b", line, re.IGNORECASE) else "미사용"
                location = "근무지내" if "근무지내" in line or "관내" in line else ""
                rows.append([name_candidates[0], time_match.group(), location, vehicle])
    finally:
        document.close()
    if not rows:
        raise ValueError("스캔 PDF에서 출장 행을 찾지 못했습니다. Tesseract 한국어 언어팩과 PDF 형식을 확인해 주세요.")
    return ["성명", "출장시간", "근무지", "차량사용여부", "소속", "직급"], rows


def coordinate_pdf_rows(path: Path) -> tuple[list[str], list[list[str]]]:
    """인사랑 PDF의 순번 구간을 기준으로 페이지를 넘는 행까지 복원한다."""
    import pdfplumber

    rows: list[list[str]] = []
    with pdfplumber.open(path) as pdf:
        words: list[dict[str, Any]] = []
        for page in pdf.pages:
            words.extend(page.extract_words())
        words.sort(key=lambda word: float(word["doctop"]))
        serials = [
            word for word in words
            if 30 <= float(word["x0"]) <= 50 and re.fullmatch(r"\d+", word["text"].strip())
        ]
        serials.sort(key=lambda word: float(word["doctop"]))
        previous_name = ""
        for index, serial in enumerate(serials):
            start = float(serial["doctop"]) - 3
            end = float(serials[index + 1]["doctop"]) - 1 if index + 1 < len(serials) else float("inf")
            group = [word for word in words if start <= float(word["doctop"]) <= end]
            name_parts = [
                word for word in group
                if 125 <= float(word["x0"]) <= 165
                and re.fullmatch(r"[가-힣\s]+", word["text"])
                and word["text"] != "성명"
                and float(word["doctop"]) <= float(serial["doctop"]) + 5
            ]
            if name_parts:
                name_top = max(float(word["doctop"]) for word in name_parts)
                name_parts = [word for word in name_parts if name_top - 2 <= float(word["doctop"]) <= name_top + 2]
            name_parts.sort(key=lambda word: float(word["x0"]))
            display_name = re.sub(r"[\s\u200b]+", "", "".join(word["text"] for word in name_parts))
            if display_name:
                previous_name = display_name
            else:
                display_name = previous_name
            duration_candidates = [
                word["text"] for word in group
                if 235 <= float(word["x0"]) <= 305 and re.search(r"\d", word["text"])
            ]
            duration_words = [
                word for word in duration_candidates
                if any(unit in word for unit in ("시간", "분", "일"))
            ] or duration_candidates
            location = "".join(word["text"] for word in group if 295 <= float(word["x0"]) <= 335)
            vehicle_words = [word["text"] for word in group if 325 <= float(word["x0"]) <= 380]
            context_words = [word["text"] for word in group if 50 <= float(word["x0"]) <= 130]
            rank = next((display_rank(word) for word in context_words if any(key in word for key in ("사무관", "주사", "서기"))), "")
            department = " ".join(word for word in context_words if display_rank(word) != rank and word not in {"(", ")"})
            if display_name and duration_words and vehicle_words and ("지내" in location or "지외" in location):
                duration = duration_words[0]
                rows.append([display_name, duration, location, vehicle_words[0], department, rank, duration])
    return ["성명", "출장시간", "근무지", "차량사용여부", "소속", "직급", "출장일수"], rows


def load_rows(path: Path) -> tuple[list[str], list[list[Any]]]:
    """Excel, CSV, 텍스트형 PDF를 표로 읽는다."""
    suffix = path.suffix.lower()
    if suffix in {".xls", ".xlsx", ".xlsm", ".csv"}:
        import pandas as pd

        if suffix == ".csv":
            frame = pd.read_csv(path, dtype=object, encoding="utf-8-sig")
        else:
            frame = pd.read_excel(path, dtype=object, sheet_name=0, header=None)
        frame = frame.dropna(how="all").fillna("")
        raw = frame.values.tolist()
        if not raw:
            return [], []
        # 인사랑 파일은 제목 행이 앞에 올 수 있으므로 헤더 후보를 자동 탐색한다.
        header_row = next(
            (index for index, row in enumerate(raw[:20]) if any(header_score(clean_text(cell), PERSON_ALIASES) for cell in row)),
            0,
        )
        return [clean_text(cell) or f"열{index + 1}" for index, cell in enumerate(raw[header_row])], raw[header_row + 1 :]
    if suffix == ".pdf":
        import pdfplumber

        coordinate_headers, coordinate_rows = coordinate_pdf_rows(path)
        if coordinate_rows:
            return coordinate_headers, coordinate_rows
        rows: list[list[Any]] = []
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                rows.extend(row for table in (page.extract_tables() or []) for row in table if row)
        if not rows:
            return ocr_pdf_rows(path)
        header_row = next(
            (index for index, row in enumerate(rows[:20]) if any(header_score(clean_text(cell), PERSON_ALIASES) for cell in row)),
            0,
        )
        return [clean_text(cell) or f"열{index + 1}" for index, cell in enumerate(rows[header_row])], rows[header_row + 1 :]
    raise ValueError("지원 형식은 .xls, .xlsx, .xlsm, .csv, .pdf 입니다.")


def summarize(headers: list[str], rows: list[list[Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    if choose_column(headers, TIME_ALIASES) is not None and choose_column(headers, VEHICLE_ALIASES) is not None:
        return summarize_trip_rows(headers, rows)
    person_index = choose_column(headers, PERSON_ALIASES)
    if person_index is None:
        raise ValueError("사람 이름 열을 찾지 못했습니다. 열 이름에 '성명' 또는 '출장자'가 있어야 합니다.")
    amount_index = choose_column(headers, AMOUNT_ALIASES)
    if amount_index == person_index:
        amount_index = None
    if amount_index is None:
        candidates = [
            index
            for index, header in enumerate(headers)
            if header_score(header, COST_ALIASES)
            and not any(word in clean_text(header) for word in IGNORE_AMOUNT_WORDS)
        ]
    else:
        candidates = [amount_index]

    detail: list[dict[str, Any]] = []
    totals: dict[str, dict[str, Any]] = defaultdict(lambda: {"출장 횟수": 0, "총 출장비": 0.0})
    previous_person = ""
    for row in rows:
        values = list(row) + [""] * max(0, len(headers) - len(row))
        person = clean_text(values[person_index])
        if not person:
            person = previous_person
        if not person or person in {"합계", "총계"}:
            continue
        previous_person = person
        amount = sum(number(values[index]) for index in candidates if index < len(values))
        record = {"성명": person, "출장비": round(amount, 2)}
        detail.append(record)
        totals[person]["출장 횟수"] += 1
        totals[person]["총 출장비"] += amount
    summary = [{"성명": person, "출장 횟수": data["출장 횟수"], "총 출장비": round(data["총 출장비"], 2)} for person, data in totals.items()]
    return detail, summary


def summarize_mapped_detail(detail: list[dict[str, Any]]) -> list[dict[str, Any]]:
    totals: dict[str, dict[str, Any]] = {}
    for row in detail:
        person = row["성명"]
        if person not in totals:
            totals[person] = {"소속": row.get("소속", ""), "직급": row.get("직급", ""), "출장 횟수": 0, "관내 출장 수": 0, "관외 총일수": 0, "관내 차량 수": 0, "관외 차량 수": 0, "4시간 미만 수": 0, "총 출장비": 0}
        totals[person]["출장 횟수"] += 1
        totals[person]["총 출장비"] += int(row["적용금액"])
        if row.get("관외여부"):
            totals[person]["관외 총일수"] += int(row.get("출장일수", 1))
            totals[person]["관외 차량 수"] += int(vehicle_used(row["차량사용여부"]))
        else:
            totals[person]["관내 출장 수"] += 1
            totals[person]["관내 차량 수"] += int(vehicle_used(row["차량사용여부"]))
            totals[person]["4시간 미만 수"] += int((duration_hours(row["출장시간"]) or 0) < 4)
    return [{"성명": person, **data} for person, data in totals.items()]


def order_summary(summary: list[dict[str, Any]], order_text: str) -> list[dict[str, Any]]:
    """사무분장표 순서를 우선 적용하고, 목록에 없는 사람은 기존 순서로 뒤에 둔다."""
    requested = [
        re.sub(r"\s+", "", line.strip())
        for line in order_text.splitlines()
        if line.strip()
    ]
    by_name = {row["성명"]: row for row in summary}
    ordered: list[dict[str, Any]] = []
    used: set[str] = set()
    for name in requested:
        if name in by_name and name not in used:
            ordered.append(by_name[name])
            used.add(name)
    ordered.extend(row for row in summary if row["성명"] not in used)
    return ordered


def save_standard_result(output: Path, summary: list[dict[str, Any]], detail: list[dict[str, Any]], template: Path) -> None:
    """시 표준 지급명세서의 관내 시트 구성으로 XLSX를 생성한다."""
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    title = "여  비  지  급  명 세  서(2026. 7월분)"
    try:
        import xlrd

        source_book = xlrd.open_workbook(template, on_demand=True)
        source_sheet = source_book.sheet_by_name("관내")
        if source_sheet.cell_value(0, 1):
            title = str(source_sheet.cell_value(0, 1))
    except Exception:
        pass

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "관내"
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.merge_cells("B1:Y1")
    sheet["B1"] = title
    sheet["B1"].font = Font(size=14, bold=True)
    sheet["B1"].alignment = center
    sheet.merge_cells("B3:Y3")
    sheet["B3"] = "* 화성시 재무회계 규칙 제21조5항 예산집행품의 생략[직무수행경비,공공요금,제세공과금,인건비,여비]"
    sheet["B4"] = "(별지서식 제47호 서식)"
    merges = ("B5:D5", "E5:E6", "F5:H5", "I5:K5", "L5:M5", "N5:O5", "P5:P6", "Q5:Q6", "R5:R6", "S5:T5", "U5:V5", "W5:W6", "X5:X6", "Y5:Y6")
    for merge in merges:
        sheet.merge_cells(merge)
    top_headers = {
        "B5": "출장자", "E5": "월/일", "F5": "출장지", "I5": "운임",
        "L5": "관내출장현황\n(A=일+일비)", "N5": "관외출장현황\n(A=일+일비)",
        "P5": "식\n비\n(B)", "Q5": "숙\n박\n료\n(C)", "R5": "교\n통\n비\n(D)",
        "S5": "배차현황(E)", "U5": "4시간미만(F)", "W5": "계\n(A+A-1+B+C+D-E-F)",
        "X5": "비고", "Y5": "확인\n(서명란)",
    }
    sub_headers = {
        "B6": "소속", "C6": "직급", "D6": "성명", "F6": "출발", "G6": "경유", "H6": "도착",
        "I6": "종별\n(등급)", "J6": "거리", "K6": "요금", "L6": "일/야", "M6": "일비",
        "N6": "일/야", "O6": "일비", "S6": "일수", "T6": "감액\n(일비의\n50%)",
        "U6": "일수", "V6": "감액\n(일비의\n50%)",
    }
    for address, value in {**top_headers, **sub_headers}.items():
        sheet[address] = value
    for row in range(5, 7):
        for column in range(2, 26):
            cell = sheet.cell(row, column)
            cell.font = Font(bold=True)
            cell.alignment = center
            cell.border = border
    sheet.row_dimensions[5].height = 38
    sheet.row_dimensions[6].height = 58
    for row_number in range(7, max(25, 7 + len(summary))):
        for column in range(2, 26):
            sheet.cell(row_number, column).border = border
            sheet.cell(row_number, column).alignment = center
    for index, person in enumerate(summary, start=7):
        row_number = index
        sheet.cell(index, 2).value = person.get("소속", "")
        sheet.cell(index, 3).value = person.get("직급", "")
        sheet.cell(index, 4).value = person["성명"]
        sheet.cell(index, 5).value = month_day_range(detail)
        sheet.cell(index, 12).value = person["관내 출장 수"]
        sheet.cell(index, 13).value = f"=IF(L{row_number}>0,20000,0)"
        sheet.cell(index, 14).value = person["관외 총일수"]
        sheet.cell(index, 15).value = f"=IF(N{row_number}>0,25000,0)"
        sheet.cell(index, 16).value = f"=N{row_number}*25000"
        sheet.cell(index, 19).value = person["관내 차량 수"] + person["관외 차량 수"]
        sheet.cell(index, 20).value = f"={person['관내 차량 수']}*10000+{person['관외 차량 수']}*12500"
        sheet.cell(index, 21).value = person["4시간 미만 수"]
        sheet.cell(index, 22).value = f"=U{row_number}*10000"
        sheet.cell(index, 23).value = f"=L{row_number}*M{row_number}+N{row_number}*O{row_number}+P{row_number}-T{row_number}-V{row_number}"
        note = f" / {person['비고']}" if person.get("비고") else ""
        sheet.cell(index, 24).value = f"출장 {person['출장 횟수']}건{note}"
        for column in (14, 15, 16):
            sheet.cell(index, column).font = Font(color="FF0000")
        for column in (13, 15, 16, 20, 22, 23):
            sheet.cell(index, column).number_format = "#,##0"
    total_row = 25 if len(summary) <= 18 else 7 + len(summary)
    sheet.cell(total_row, 3).value = "합계"
    sheet.cell(total_row, 23).value = f"=SUM(W7:W{total_row - 1})"
    sheet.cell(total_row, 23).number_format = "#,##0"
    for column in range(2, 26):
        sheet.cell(total_row, column).border = border
        sheet.cell(total_row, column).alignment = center
        sheet.cell(total_row, column).font = Font(bold=True)
    sheet.freeze_panes = "E6"
    widths = {"B": 16, "C": 12, "D": 12, "E": 10, "F": 4, "G": 4, "H": 4, "I": 4, "J": 4, "K": 4, "L": 4, "M": 12, "N": 4, "O": 12, "P": 14, "Q": 4, "R": 4, "S": 4, "T": 14, "U": 4, "V": 14, "W": 16, "X": 24, "Y": 12}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    workbook.save(output)


def save_result(output: Path, detail: list[dict[str, Any]], summary: list[dict[str, Any]]) -> None:
    import pandas as pd

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(summary).to_excel(writer, sheet_name="사람별 집계", index=False)
        pd.DataFrame(detail).to_excel(writer, sheet_name="출장별 내역", index=False)
        workbook = writer.book
        for worksheet in workbook.worksheets:
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            for column in worksheet.columns:
                width = min(max(max(len(str(cell.value or "")) for cell in column) + 2, 10), 30)
                worksheet.column_dimensions[column[0].column_letter].width = width
        summary_sheet = workbook["사람별 집계"]
        summary_sheet.append(["전체", sum(row["출장 횟수"] for row in summary), sum(row["총 출장비"] for row in summary)])
        for cell in summary_sheet[1]:
            font = copy(cell.font)
            font.bold = True
            cell.font = font


def run(source: str, destination: str | None = None) -> Path:
    source_path = Path(source).expanduser().resolve()
    headers, rows = load_rows(source_path)
    detail, summary = summarize(headers, rows)
    if not summary:
        raise ValueError("집계할 출장 행이 없습니다.")
    output = Path(destination).expanduser().resolve() if destination else source_path.with_name(f"{source_path.stem}_관내출장_집계.xlsx")
    save_result(output, detail, summary)
    return output


def web_app() -> None:
    import streamlit as st

    st.set_page_config(page_title="관내출장비 계산기", page_icon="💼", layout="wide")
    st.markdown(
        """
        <style>
        .block-container { max-width: 1100px; padding-top: 3rem; }
        [data-testid="stMetricValue"] { color: #0f766e; }
        </style>
        """,
        unsafe_allow_html=True,
    )
    st.title("관내출장비 계산기")
    st.caption("인사랑 결재내역 PDF를 올리면 출장시간과 차량 사용 여부를 기준으로 자동 계산합니다.")
    st.info("관내: 4시간 미만 1만원 · 4시간 이상 2만원 · 차량 사용 시 1만원 차감 / 관외: 일비·식비 각 1일 2만5천원 · 관용차 사용 시 1일 1만2천5백원 감액")
    duty_order_text = st.text_area(
        "사무분장표 순서",
        height=140,
        placeholder="이름1\n이름2\n이름3\n...",
        help="실제 지급명세서에 표시할 순서대로 실제 이름을 한 줄씩 입력하세요.",
    )
    uploaded = st.file_uploader("결재내역 PDF 업로드", type=["pdf"], accept_multiple_files=False)
    if uploaded is None:
        st.write("PDF 파일을 업로드하면 사람별 집계 결과가 여기에 표시됩니다.")
        return

    source_path: Path | None = None
    output_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as source_file:
            source_file.write(uploaded.getvalue())
            source_path = Path(source_file.name)
        headers, rows = load_rows(source_path)
        detail, summary = summarize(headers, rows)
        summary = order_summary(summary, duty_order_text)
        if not summary:
            raise ValueError("집계할 출장 행이 없습니다.")
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as output_file:
            output_path = Path(output_file.name)
        template = Path(__file__).with_name("7월 여비지급명세서.xls")
        save_standard_result(output_path, summary, detail, template)
    except Exception as error:
        st.error(str(error))
        st.caption("스캔 PDF라면 Tesseract OCR 한국어 언어팩이 설치되어 있어야 합니다.")
        return
    finally:
        if source_path:
            source_path.unlink(missing_ok=True)

    total_count = sum(row["출장 횟수"] for row in summary)
    total_amount = sum(row["총 출장비"] for row in summary)
    first, second, third = st.columns(3)
    first.metric("전체 출장 건수", f"{total_count:,}건")
    second.metric("전체 출장비", f"{total_amount:,}원")
    third.metric("출장자 수", f"{len(summary):,}명")
    st.subheader("사람별 집계")
    st.dataframe(summary, width="stretch", hide_index=True)
    st.subheader("출장별 계산 내역")
    st.dataframe(detail, width="stretch", hide_index=True)
    if output_path:
        st.download_button(
            "엑셀 다운로드",
            data=output_path.read_bytes(),
            file_name="7월 여비지급명세서_완성.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
        )
        output_path.unlink(missing_ok=True)


if __name__ == "__main__":
    web_app()